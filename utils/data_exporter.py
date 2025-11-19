# -*- coding: utf-8 -*-
"""
数据导出模块
支持导出为Excel和CSV格式
"""

import os
import csv
from typing import List, Dict, Optional
from datetime import datetime
from utils.logger import get_logger

logger = get_logger()


class DataExporter:
    """数据导出器"""
    
    def __init__(self):
        self.logger = logger
    
    def export_to_csv(self, records: List[Dict], user_name: str, output_dir: str = None) -> Optional[str]:
        """导出成绩记录为CSV格式
        
        Args:
            records: 成绩记录列表
            user_name: 用户名称
            output_dir: 输出目录，默认为当前目录
            
        Returns:
            导出文件的路径，失败返回None
        """
        try:
            if not records:
                logger.warning('没有可导出的数据')
                return None
            
            # 确定输出目录
            if output_dir is None:
                output_dir = os.path.expanduser('~/Desktop')
            
            # 确保目录存在
            os.makedirs(output_dir, exist_ok=True)
            
            # 生成文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'{user_name}_成绩记录_{timestamp}.csv'
            filepath = os.path.join(output_dir, filename)
            
            logger.info(f'开始导出CSV: {filepath}')
            
            # 写入CSV
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
                # 定义表头
                fieldnames = ['日期', '必选项目', '必选成绩', '必选得分',
                            '第一类选考项目', '第一类成绩', '第一类得分',
                            '第二类选考项目', '第二类成绩', '第二类得分',
                            '总分', '等级']
                
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                # 写入数据
                for record in records:
                    row = self._format_record_for_csv(record)
                    writer.writerow(row)
            
            logger.info(f'成功导出 {len(records)} 条记录到: {filepath}')
            return filepath
            
        except Exception as e:
            logger.error(f'导出CSV失败: {e}', exc_info=True)
            return None
    
    def export_to_excel(self, records: List[Dict], user_name: str, output_dir: str = None) -> Optional[str]:
        """导出成绩记录为Excel格式
        
        Args:
            records: 成绩记录列表
            user_name: 用户名称
            output_dir: 输出目录，默认为桌面
            
        Returns:
            导出文件的路径，失败返回None
        """
        try:
            # 延迟导入，如果没有安装openpyxl也不影响CSV导出
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            except ImportError:
                logger.error('未安装openpyxl库，无法导出Excel格式')
                return None
            
            if not records:
                logger.warning('没有可导出的数据')
                return None
            
            # 确定输出目录
            if output_dir is None:
                output_dir = os.path.expanduser('~/Desktop')
            
            # 确保目录存在
            os.makedirs(output_dir, exist_ok=True)
            
            # 生成文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'{user_name}_成绩记录_{timestamp}.xlsx'
            filepath = os.path.join(output_dir, filename)
            
            logger.info(f'开始导出Excel: {filepath}')
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = '成绩记录'
            
            # 设置表头样式
            header_font = Font(bold=True, size=12, color='FFFFFF')
            header_fill = PatternFill(start_color='16a085', end_color='16a085', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # 定义表头
            headers = ['日期', '必选项目', '必选成绩', '必选得分',
                      '第一类选考项目', '第一类成绩', '第一类得分',
                      '第二类选考项目', '第二类成绩', '第二类得分',
                      '总分', '等级']
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 设置数据样式
            data_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入数据
            for row_idx, record in enumerate(records, 2):
                formatted_record = self._format_record_for_excel(record)
                for col_idx, value in enumerate(formatted_record, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = data_alignment
                    cell.border = border
                    
                    # 根据等级设置颜色
                    if col_idx == len(formatted_record):  # 等级列
                        if value == '优秀':
                            cell.fill = PatternFill(start_color='d5f4e6', end_color='d5f4e6', fill_type='solid')
                        elif value == '良好':
                            cell.fill = PatternFill(start_color='e3f2fd', end_color='e3f2fd', fill_type='solid')
                        elif value == '不及格':
                            cell.fill = PatternFill(start_color='fadbd8', end_color='fadbd8', fill_type='solid')
            
            # 调整列宽
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + col)].width = 15
            
            # 保存文件
            wb.save(filepath)
            
            logger.info(f'成功导出 {len(records)} 条记录到: {filepath}')
            return filepath
            
        except Exception as e:
            logger.error(f'导出Excel失败: {e}', exc_info=True)
            return None
    
    def _format_record_for_csv(self, record: Dict) -> Dict:
        """格式化记录用于CSV导出"""
        from config.constants import PROJECT_NAMES
        from services.score_calculator import ScoreCalculator
        
        calculator = ScoreCalculator()
        
        # 提取必选项
        required = record.get('required', {})
        req_project = list(required.keys())[0] if required else ''
        req_value = list(required.values())[0] if required else ''
        
        # 提取第一类选考
        category1 = record.get('category1', {})
        cat1_project = list(category1.keys())[0] if category1 else ''
        cat1_value = list(category1.values())[0] if category1 else ''
        
        # 提取第二类选考
        category2 = record.get('category2', {})
        cat2_project = list(category2.keys())[0] if category2 else ''
        cat2_value = list(category2.values())[0] if category2 else ''
        
        # 获取得分
        scores = record.get('scores', {})
        total_score = record.get('total_score', 0)
        grade = calculator.get_grade_level(total_score)
        
        return {
            '日期': record.get('date', ''),
            '必选项目': PROJECT_NAMES.get(req_project, req_project),
            '必选成绩': self._format_performance(req_project, req_value),
            '必选得分': f"{scores.get('required', 0):.1f}",
            '第一类选考项目': PROJECT_NAMES.get(cat1_project, cat1_project),
            '第一类成绩': self._format_performance(cat1_project, cat1_value),
            '第一类得分': f"{scores.get('category1', 0):.1f}",
            '第二类选考项目': PROJECT_NAMES.get(cat2_project, cat2_project),
            '第二类成绩': self._format_performance(cat2_project, cat2_value),
            '第二类得分': f"{scores.get('category2', 0):.1f}",
            '总分': f"{total_score:.1f}",
            '等级': grade
        }
    
    def _format_record_for_excel(self, record: Dict) -> List:
        """格式化记录用于Excel导出"""
        formatted = self._format_record_for_csv(record)
        return list(formatted.values())
    
    def _format_performance(self, project: str, value) -> str:
        """格式化成绩显示"""
        if not value:
            return ''
        
        # 时间类项目（秒转分秒）
        if project in ['1000m', '800m']:
            minutes = int(value) // 60
            seconds = int(value) % 60
            return f"{minutes}'{seconds}\""
        # 其他项目直接显示
        else:
            return str(value)
